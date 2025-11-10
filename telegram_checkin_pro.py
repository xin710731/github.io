# telegram_checkin_pro_v3.py
# 完整版（aiogram 3）：多语言（中文 / English / Bahasa Indonesia）
# - 打卡 / 休息（Emoji识别）/ 回座统计 / 超时提醒
# - 多管理员设置面板（多 ID）
# - 管理员日志（写入 admin_logs）
# - 自动/手动 报表（Excel .xlsx，中文文件名，带群名）
# - 自动在首次使用时为群插入 settings 初始行
#
# 依赖:
# pip install aiogram==3.1.0 aiosqlite python-dotenv openpyxl apscheduler

import asyncio
import aiosqlite
import io
import os
import re
import logging
from datetime import datetime, timedelta, date, time
from typing import Optional, Dict, List

from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

# ---------------------------
# 配置
# ---------------------------
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").replace(" ", "").split(",") if x]

if not BOT_TOKEN:
    raise RuntimeError("请在 .env 中设置 BOT_TOKEN")

DB_PATH = "checkin_pro.db"
LOCAL_OFFSET = timedelta(hours=7)   # 印尼时区，可改
DAILY_REPORT_HOUR = 10
WEEKLY_REPORT_DAY = 0
WEEKLY_REPORT_HOUR = 10
MONTHLY_REPORT_DAY = 1
MONTHLY_REPORT_HOUR = 10
OVERTIME_REMINDER_INTERVAL = 3  # 分钟

BREAK_LIMITS = {
    "toilet_small": 5,
    "toilet_big": 10,
    "smoke": 5,
    "meal": 30,
}

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()
pending_media_for_chat: Dict[int, str] = {}  # chat_id -> state string

# ---------------------------
# 多语言字典
# ---------------------------
LANG_TEXT = {
    "zh": {
        "welcome": "欢迎使用打卡机器人。请通过菜单进行操作。",
        "menu": [
            ["🏁 上班打卡", "🏁 下班签退"],
            ["🚶‍♂️ 小厕开始", "🚽 大厕开始"],
            ["🚬 抽烟开始", "🍱 吃饭开始"],
            ["💺 回座", "📊 今日统计"],
            ["📈 排行榜", "⚙️ 设置"]
        ],
        "start_work": "✅ 上班打卡成功！",
        "end_work": "🕒 下班签退成功。",
        "return_seat": "💼 欢迎回来！",
        "no_break_running": "💼 欢迎回来！",
        "today_title": "📋 <b>当日工作总结</b>",
        "today_user": "👤 用户",
        "total_work": "• 工作总计",
        "total_break": "• 休息时间",
        "leave_times": "• 离开次数",
        "meal": "🍱 吃饭",
        "toilet": "🚻 厕所",
        "smoke": "🚬 抽烟",
        "leaderboard_title": "🏆 本群今日排行榜",
        "no_data": "暂无数据。",
        "admin_only": "🚫 你不是管理员，无法访问设置菜单。",
        "admin_menu_title": "⚙️ 管理员设置菜单：",
        "adm_set_text": "📝 设置提醒文字",
        "adm_set_media": "🖼️ 上传提醒图片",
        "adm_toggle_weekly": "📅 切换周报",
        "adm_toggle_monthly": "🗓️ 切换月报",
        "adm_reset_leaderboard": "🔄 重置排行榜",
        "adm_send_daily": "📤 手动发送日报",
        "enter_new_text": "请输入新的提醒文字（发送一条消息即可）：",
        "text_updated": "✅ 提醒文字已更新。",
        "send_image": "请发送一张图片作为提醒媒体：",
        "image_updated": "✅ 提醒图片已更新。",
        "no_permission": "没有权限",
        "weekly_on": "📅 周报功能 ✅ 已开启",
        "weekly_off": "📅 周报功能 ❌ 已关闭",
        "monthly_on": "🗓️ 月报功能 ✅ 已开启",
        "monthly_off": "🗓️ 月报功能 ❌ 已关闭",
        "done": "操作完成 ✅",
        "reset_done": "🔄 排行榜已重置！",
        "daily_sent": "📊 日报已发送给管理员。",
        "manual_daily_done": "✅ 日报已生成并发送给管理员。",
        "stats_error": "❌ 统计出错",
        "reminder_default": "你已开始 {label} ，预计 {limit} 分钟。",
        "overtime_default": "⚠️ <a href='tg://user?id={uid}'>你</a> 已超时，请尽快回座。",
        "tz_label": "时区",
        "not_admin": "🚫 你不是管理员，无权执行此操作。",
    },
    "en": {
        "welcome": "Welcome! Please use the menu to operate.",
        "menu": [
            ["🏁 Clock In", "🏁 Clock Out"],
            ["🚶‍♂️ Small Toilet", "🚽 Big Toilet"],
            ["🚬 Smoke Break", "🍱 Meal Break"],
            ["💺 Back to Seat", "📊 Today Summary"],
            ["📈 Leaderboard", "⚙️ Settings"]
        ],
        "start_work": "✅ Clock-in successful!",
        "end_work": "🕒 Clock-out successful.",
        "return_seat": "💺 Welcome back!",
        "no_break_running": "💺 Welcome back!",
        "today_title": "📋 <b>Today's Summary</b>",
        "today_user": "👤 User",
        "total_work": "• Work Total",
        "total_break": "• Break Time",
        "leave_times": "• Leaves",
        "meal": "🍱 Meal",
        "toilet": "🚻 Toilet",
        "smoke": "🚬 Smoke",
        "leaderboard_title": "🏆 Today's Leaderboard",
        "no_data": "No data.",
        "admin_only": "🚫 You are not an admin.",
        "admin_menu_title": "⚙️ Admin Settings:",
        "adm_set_text": "📝 Set Reminder Text",
        "adm_set_media": "🖼️ Upload Reminder Image",
        "adm_toggle_weekly": "📅 Toggle Weekly Report",
        "adm_toggle_monthly": "🗓️ Toggle Monthly Report",
        "adm_reset_leaderboard": "🔄 Reset Leaderboard",
        "adm_send_daily": "📤 Send Daily Report Now",
        "enter_new_text": "Please send the new reminder text (one message):",
        "text_updated": "✅ Reminder text updated.",
        "send_image": "Please send an image as the reminder media:",
        "image_updated": "✅ Reminder image updated.",
        "no_permission": "No permission",
        "weekly_on": "📅 Weekly report ✅ ON",
        "weekly_off": "📅 Weekly report ❌ OFF",
        "monthly_on": "🗓️ Monthly report ✅ ON",
        "monthly_off": "🗓️ Monthly report ❌ OFF",
        "done": "Done ✅",
        "reset_done": "🔄 Leaderboard reset!",
        "daily_sent": "📊 Daily report has been sent to admins.",
        "manual_daily_done": "✅ Daily reports generated and sent to admins.",
        "stats_error": "❌ Stats error",
        "reminder_default": "You started {label}, expected {limit} minutes.",
        "overtime_default": "⚠️ <a href='tg://user?id={uid}'>You</a> exceeded the limit, please return.",
        "tz_label": "Timezone",
        "not_admin": "🚫 You are not an admin.",
    },
    "id": {
        "welcome": "Selamat datang! Silakan gunakan menu untuk beroperasi.",
        "menu": [
            ["🏁 Masuk Kerja", "🏁 Pulang Kerja"],
            ["🚶‍♂️ Toilet Kecil", "🚽 Toilet Besar"],
            ["🚬 Istirahat Merokok", "🍱 Istirahat Makan"],
            ["💺 Kembali Duduk", "📊 Ringkasan Hari Ini"],
            ["📈 Papan Peringkat", "⚙️ Pengaturan"]
        ],
        "start_work": "✅ Masuk kerja tercatat!",
        "end_work": "🕒 Pulang kerja tercatat.",
        "return_seat": "💺 Selamat datang kembali!",
        "no_break_running": "💺 Selamat datang kembali!",
        "today_title": "📋 <b>Ringkasan Hari Ini</b>",
        "today_user": "👤 Pengguna",
        "total_work": "• Total Kerja",
        "total_break": "• Waktu Istirahat",
        "leave_times": "• Jumlah Keluar",
        "meal": "🍱 Makan",
        "toilet": "🚻 Toilet",
        "smoke": "🚬 Merokok",
        "leaderboard_title": "🏆 Papan Peringkat Hari Ini",
        "no_data": "Tidak ada data.",
        "admin_only": "🚫 Anda bukan admin.",
        "admin_menu_title": "⚙️ Pengaturan Admin:",
        "adm_set_text": "📝 Atur Teks Pengingat",
        "adm_set_media": "🖼️ Unggah Gambar Pengingat",
        "adm_toggle_weekly": "📅 Alihkan Laporan Mingguan",
        "adm_toggle_monthly": "🗓️ Alihkan Laporan Bulanan",
        "adm_reset_leaderboard": "🔄 Setel Ulang Papan Peringkat",
        "adm_send_daily": "📤 Kirim Laporan Harian Sekarang",
        "enter_new_text": "Kirim teks pengingat baru (satu pesan):",
        "text_updated": "✅ Teks pengingat diperbarui.",
        "send_image": "Kirim gambar sebagai media pengingat:",
        "image_updated": "✅ Gambar pengingat diperbarui.",
        "no_permission": "Tidak ada izin",
        "weekly_on": "📅 Laporan mingguan ✅ AKTIF",
        "weekly_off": "📅 Laporan mingguan ❌ NONAKTIF",
        "monthly_on": "🗓️ Laporan bulanan ✅ AKTIF",
        "monthly_off": "🗓️ Laporan bulanan ❌ NONAKTIF",
        "done": "Selesai ✅",
        "reset_done": "🔄 Papan peringkat direset!",
        "daily_sent": "📊 Laporan harian telah dikirim ke admin.",
        "manual_daily_done": "✅ Laporan harian dibuat dan dikirim ke admin.",
        "stats_error": "❌ Kesalahan statistik",
        "reminder_default": "Anda memulai {label}, perkiraan {limit} menit.",
        "overtime_default": "⚠️ <a href='tg://user?id={uid}'>Anda</a> melewati batas, harap kembali.",
        "tz_label": "Zona waktu",
        "not_admin": "🚫 Anda bukan admin.",
    }
}

# 供按钮匹配的三语文本集合
def _collect_menu_keys() -> Dict[str, List[str]]:
    # 按菜单位置收集（稳定）
    keys = {
        "start_work": [],
        "end_work": [],
        "return_seat": [],
        "today_summary": [],
        "leaderboard": [],
        "settings": [],
    }
    for lang in ("zh", "en", "id"):
        m = LANG_TEXT[lang]["menu"]
        keys["start_work"].append(m[0][0])
        keys["end_work"].append(m[0][1])
        keys["return_seat"].append(m[3][0])
        keys["today_summary"].append(m[3][1])
        keys["leaderboard"].append(m[4][0])
        keys["settings"].append(m[4][1])
    return keys

MENU_KEYS = _collect_menu_keys()

# ---------------------------
# 时间与格式工具
# ---------------------------
def now_utc() -> datetime:
    return datetime.utcnow()

def now_local() -> datetime:
    return datetime.utcnow() + LOCAL_OFFSET

def to_str(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def parse_str(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except:
        return None

def fmt_hm_local(dt_utc: Optional[datetime]) -> str:
    if not dt_utc:
        return "-"
    return (dt_utc + LOCAL_OFFSET).strftime("%H:%M")

def today_local_date() -> date:
    return (datetime.utcnow() + LOCAL_OFFSET).date()

def minutes_between(a: Optional[datetime], b: Optional[datetime]) -> int:
    if not a or not b:
        return 0
    return max(0, int((b - a).total_seconds() // 60))

def fmt_minutes(m: int) -> str:
    if m >= 60:
        h = m // 60
        mm = m % 60
        return f"{h}小时{mm}分钟"
    return f"{m}分钟"

# ---------------------------
# 语言与菜单工具
# ---------------------------
def detect_lang(user: types.User) -> str:
    lang = (user.language_code or "").lower()
    if lang.startswith("zh"):
        return "zh"
    elif lang.startswith("id"):
        return "id"
    else:
        return "en"

def get_menu(lang="zh"):
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=txt) for txt in row] for row in LANG_TEXT[lang]["menu"]],
        resize_keyboard=True
    )

def text_in_keys(text: str, key: str) -> bool:
    return text in MENU_KEYS[key]

# ---------------------------
# DB 初始化（含 admin_logs）
# ---------------------------
async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS work_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                chat_id INTEGER,
                start_time TEXT,
                end_time TEXT
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS break_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                chat_id INTEGER,
                type TEXT,
                start_time TEXT,
                end_time TEXT
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                chat_id INTEGER PRIMARY KEY,
                reminder_text TEXT,
                reminder_media_file_id TEXT,
                weekly_report_enabled INTEGER DEFAULT 0,
                monthly_report_enabled INTEGER DEFAULT 0
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS admin_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER,
                admin_id INTEGER,
                action TEXT,
                details TEXT,
                created_at TEXT
            )
        """)
        await db.commit()
    logger.info("数据库初始化完成。")

# ---------------------------
# 设置/日志辅助
# ---------------------------
async def ensure_settings(chat_id: int):
    """确保 settings 表存在该 chat_id 的行（首次使用自动插入）"""
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT 1 FROM settings WHERE chat_id = ?", (chat_id,))
        found = await cur.fetchone()
        if not found:
            await db.execute(
                "INSERT INTO settings (chat_id, reminder_text, reminder_media_file_id, weekly_report_enabled, monthly_report_enabled) VALUES (?, ?, ?, 0, 0)",
                (chat_id, None, None)
            )
            await db.commit()

async def set_chat_setting(chat_id: int, key: str, value):
    await ensure_settings(chat_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE settings SET {key} = ? WHERE chat_id = ?", (value, chat_id))
        await db.commit()

async def get_chat_settings(chat_id: int):
    await ensure_settings(chat_id)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT reminder_text, reminder_media_file_id, weekly_report_enabled, monthly_report_enabled FROM settings WHERE chat_id = ?", (chat_id,))
        row = await cur.fetchone()
    return {"reminder_text": row[0], "reminder_media_file_id": row[1], "weekly_report_enabled": row[2], "monthly_report_enabled": row[3]}

async def get_chats_with_setting_enabled(col_name: str):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(f"SELECT chat_id FROM settings WHERE {col_name} = 1")
        rows = await cur.fetchall()
    return [r[0] for r in rows]

async def log_admin_action(chat_id: int, admin_id: int, action: str, details: str = ""):
    created_at = to_str(now_utc())
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO admin_logs (chat_id, admin_id, action, details, created_at) VALUES (?, ?, ?, ?, ?)",
            (chat_id, admin_id, action, details, created_at)
        )
        await db.commit()

# ---------------------------
# 打卡 / 休息 数据写入（均确保 settings 存在）
# ---------------------------
async def start_work(user_id: int, chat_id: int):
    await ensure_settings(chat_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("INSERT INTO work_sessions (user_id, chat_id, start_time) VALUES (?, ?, ?)",
                         (user_id, chat_id, to_str(now_utc())))
        await db.commit()

async def end_work(user_id: int, chat_id: int):
    await ensure_settings(chat_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE work_sessions SET end_time = ? WHERE user_id=? AND chat_id=? AND end_time IS NULL",
                         (to_str(now_utc()), user_id, chat_id))
        await db.commit()

async def start_break(user_id: int, chat_id: int, btype: str):
    await ensure_settings(chat_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("INSERT INTO break_sessions (user_id, chat_id, type, start_time) VALUES (?, ?, ?, ?)",
                         (user_id, chat_id, btype, to_str(now_utc())))
        await db.commit()

async def end_break(user_id: int, chat_id: int):
    await ensure_settings(chat_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE break_sessions SET end_time = ? WHERE user_id=? AND chat_id=? AND end_time IS NULL",
                         (to_str(now_utc()), user_id, chat_id))
        await db.commit()

# ---------------------------
# 菜单
# ---------------------------
def get_admin_menu(lang: str):
    t = LANG_TEXT[lang]
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t["adm_set_text"], callback_data="admin:set_text")],
        [InlineKeyboardButton(text=t["adm_set_media"], callback_data="admin:set_media")],
        [InlineKeyboardButton(text=t["adm_toggle_weekly"], callback_data="admin:toggle_weekly")],
        [InlineKeyboardButton(text=t["adm_toggle_monthly"], callback_data="admin:toggle_monthly")],
        [InlineKeyboardButton(text=t["adm_reset_leaderboard"], callback_data="admin:reset_leaderboard")],
        [InlineKeyboardButton(text=t["adm_send_daily"], callback_data="admin:send_daily_report")]
    ])

# ---------------------------
# Handlers: 基本交互
# ---------------------------
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    lang = detect_lang(message.from_user)
    await ensure_settings(message.chat.id)
    await message.reply(LANG_TEXT[lang]["welcome"], reply_markup=get_menu(lang))

@dp.message(F.text.func(lambda s: text_in_keys(s, "start_work")))
async def handler_start_work(message: types.Message):
    lang = detect_lang(message.from_user)
    await start_work(message.from_user.id, message.chat.id)
    await message.reply(f"{LANG_TEXT[lang]['start_work']} ({fmt_hm_local(now_utc())})", reply_markup=get_menu(lang))

@dp.message(F.text.func(lambda s: text_in_keys(s, "end_work")))
async def handler_end_work(message: types.Message):
    lang = detect_lang(message.from_user)
    await end_work(message.from_user.id, message.chat.id)
    await message.reply(f"{LANG_TEXT[lang]['end_work']} ({fmt_hm_local(now_utc())})", reply_markup=get_menu(lang))

# 休息开始（Emoji识别：🚶, 🚽, 🚬, 🍱）
def detect_break_type_by_emoji(text: str) -> Optional[str]:
    if "🚽" in text:
        return "toilet_big"
    if "🚶" in text:
        return "toilet_small"
    if "🚬" in text:
        return "smoke"
    if "🍱" in text:
        return "meal"
    return None

def human_break_label(btype: str, lang: str) -> str:
    if btype == "meal":
        return LANG_TEXT[lang]["meal"]
    if btype in ("toilet_small", "toilet_big"):
        return LANG_TEXT[lang]["toilet"]
    if btype == "smoke":
        return LANG_TEXT[lang]["smoke"]
    return btype

@dp.message(F.text.func(lambda s: any(e in s for e in ("🚶", "🚽", "🚬", "🍱"))))
async def handler_start_break(message: types.Message):
    lang = detect_lang(message.from_user)
    btype = detect_break_type_by_emoji(message.text or "")
    if not btype:
        # 未识别则忽略
        return
    await start_break(message.from_user.id, message.chat.id, btype)
    limit = BREAK_LIMITS.get(btype, 5)
    settings = await get_chat_settings(message.chat.id)
    default_text = LANG_TEXT[lang]["reminder_default"].format(label=human_break_label(btype, lang), limit=limit)
    rtext = settings.get("reminder_text") or default_text
    await message.reply(f"{rtext}\n⏰ {fmt_hm_local(now_utc())}", reply_markup=get_menu(lang))
    asyncio.create_task(break_overtime_watcher(message.from_user.id, message.chat.id, btype, now_utc(), lang))

@dp.message(F.text.func(lambda s: text_in_keys(s, "return_seat")))
async def handler_return_seat(message: types.Message):
    lang = detect_lang(message.from_user)
    user_id = message.from_user.id
    chat_id = message.chat.id
    now = now_utc()

    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT id, type, start_time FROM break_sessions WHERE user_id=? AND chat_id=? AND end_time IS NULL ORDER BY id DESC LIMIT 1",
            (user_id, chat_id)
        )
        row = await cur.fetchone()

    if not row:
        await message.reply(f"{LANG_TEXT[lang]['no_break_running']}（{fmt_hm_local(now)}）", reply_markup=get_menu(lang))
        return

    _, btype, start_s = row
    sdt = parse_str(start_s)
    used_mins = minutes_between(sdt, now)
    human_map = {
        "zh": {"toilet_small": "小厕", "toilet_big": "大厕", "smoke": "抽烟", "meal": "吃饭"},
        "en": {"toilet_small": "small toilet", "toilet_big": "big toilet", "smoke": "smoke", "meal": "meal"},
        "id": {"toilet_small": "toilet kecil", "toilet_big": "toilet besar", "smoke": "merokok", "meal": "makan"},
    }
    human = human_map[lang].get(btype, btype)

    await end_break(user_id, chat_id)

    today = today_local_date()
    summary = await compute_daily_summary(user_id, chat_id, today)
    total_times = summary["total_leave_times"]
    total_minutes = summary["total_leave_minutes"]

    if lang == "zh":
        msg = (
            f"{LANG_TEXT[lang]['return_seat']}\n"
            f"🚶‍♂️ 本次 {human} 用时：{used_mins} 分钟\n"
            f"📅 今日累计离开 {total_times} 次，共 {fmt_minutes(total_minutes)}\n"
            f"（{fmt_hm_local(sdt)} ~ {fmt_hm_local(now)}）"
        )
    elif lang == "en":
        msg = (
            f"{LANG_TEXT[lang]['return_seat']}\n"
            f"🚶‍♂️ This {human} took: {used_mins} minutes\n"
            f"📅 Today leaves: {total_times} times, total {fmt_minutes(total_minutes)}\n"
            f"({fmt_hm_local(sdt)} ~ {fmt_hm_local(now)})"
        )
    else:  # id
        msg = (
            f"{LANG_TEXT[lang]['return_seat']}\n"
            f"🚶‍♂️ Sesi {human}: {used_mins} menit\n"
            f"📅 Hari ini keluar: {total_times} kali, total {fmt_minutes(total_minutes)}\n"
            f"({fmt_hm_local(sdt)} ~ {fmt_hm_local(now)})"
        )

    await message.reply(msg, reply_markup=get_menu(lang))

# ---------------------------
# 今日统计工具（跨天兼容）
# ---------------------------
async def get_day_intervals_for_user_in_chat(user_id: int, chat_id: int, target_date: date):
    local_start = datetime.combine(target_date, time(0, 0, 0))
    local_end = datetime.combine(target_date, time(23, 59, 59))
    utc_start = local_start - LOCAL_OFFSET
    utc_end = local_end - LOCAL_OFFSET
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT start_time, end_time FROM work_sessions "
            "WHERE user_id=? AND chat_id=? AND start_time <= ? AND (end_time IS NULL OR end_time >= ?)",
            (user_id, chat_id, to_str(utc_end), to_str(utc_start))
        )
        work_rows = await cur.fetchall()
        cur = await db.execute(
            "SELECT type, start_time, end_time FROM break_sessions "
            "WHERE user_id=? AND chat_id=? AND start_time <= ? AND (end_time IS NULL OR end_time >= ?)",
            (user_id, chat_id, to_str(utc_end), to_str(utc_start))
        )
        break_rows = await cur.fetchall()
    works = [(parse_str(s), parse_str(e) if e else None) for s, e in work_rows]
    breaks = [(t, parse_str(s), parse_str(e) if e else None) for t, s, e in break_rows]
    return works, breaks

async def compute_daily_summary(user_id: int, chat_id: int, target_date: date):
    works, breaks = await get_day_intervals_for_user_in_chat(user_id, chat_id, target_date)
    total_work = sum(minutes_between(s, e or now_utc()) for s, e in works if s)
    total_break = sum(minutes_between(s, e or now_utc()) for _, s, e in breaks if s)
    counts = {"meal": 0, "toilet_small": 0, "toilet_big": 0, "smoke": 0}
    durations = {"meal": 0, "toilet_small": 0, "toilet_big": 0, "smoke": 0}
    for btype, s, e in breaks:
        end_t = e or now_utc()
        if btype in counts:
            counts[btype] += 1
            durations[btype] += minutes_between(s, end_t)
    total_leave_times = sum(counts.values())
    total_leave_minutes = sum(durations.values())
    return {
        "total_work": total_work,
        "total_break": total_break,
        "counts": counts,
        "durations": durations,
        "total_leave_times": total_leave_times,
        "total_leave_minutes": total_leave_minutes
    }

@dp.message(F.text.func(lambda s: text_in_keys(s, "today_summary")))
async def handler_today_summary(message: types.Message):
    lang = detect_lang(message.from_user)
    user_id = message.from_user.id
    chat_id = message.chat.id
    today = today_local_date()
    try:
        summary = await compute_daily_summary(user_id, chat_id, today)
    except Exception as e:
        logger.exception("compute_daily_summary 出错")
        await message.reply(f"{LANG_TEXT[lang]['stats_error']}：{e}")
        return

    try:
        member = await bot.get_chat_member(chat_id, user_id)
        username = member.user.full_name or member.user.username or str(user_id)
    except:
        username = str(user_id)

    if lang == "zh":
        text = (
            f"{LANG_TEXT[lang]['today_title']}（{today.isoformat()}）\n"
            f"{LANG_TEXT[lang]['today_user']}：{username}\n\n"
            f"{LANG_TEXT[lang]['total_work']}：{fmt_minutes(summary['total_work'])}\n"
            f"{LANG_TEXT[lang]['total_break']}：{fmt_minutes(summary['total_break'])}\n"
            f"{LANG_TEXT[lang]['leave_times']}：{summary['total_leave_times']}\n\n"
            f"{LANG_TEXT[lang]['meal']}：{summary['counts']['meal']} 次（{fmt_minutes(summary['durations']['meal'])}）\n"
            f"{LANG_TEXT[lang]['toilet']}：{summary['counts']['toilet_small'] + summary['counts']['toilet_big']} 次（{fmt_minutes(summary['durations']['toilet_small'] + summary['durations']['toilet_big'])}）\n"
            f"{LANG_TEXT[lang]['smoke']}：{summary['counts']['smoke']} 次\n"
        )
    elif lang == "en":
        text = (
            f"{LANG_TEXT[lang]['today_title']} ({today.isoformat()})\n"
            f"{LANG_TEXT[lang]['today_user']}: {username}\n\n"
            f"{LANG_TEXT[lang]['total_work']}: {fmt_minutes(summary['total_work'])}\n"
            f"{LANG_TEXT[lang]['total_break']}: {fmt_minutes(summary['total_break'])}\n"
            f"{LANG_TEXT[lang]['leave_times']}: {summary['total_leave_times']}\n\n"
            f"{LANG_TEXT[lang]['meal']}: {summary['counts']['meal']} ({fmt_minutes(summary['durations']['meal'])})\n"
            f"{LANG_TEXT[lang]['toilet']}: {summary['counts']['toilet_small'] + summary['counts']['toilet_big']} ({fmt_minutes(summary['durations']['toilet_small'] + summary['durations']['toilet_big'])})\n"
            f"{LANG_TEXT[lang]['smoke']}: {summary['counts']['smoke']}\n"
        )
    else:
        text = (
            f"{LANG_TEXT[lang]['today_title']} ({today.isoformat()})\n"
            f"{LANG_TEXT[lang]['today_user']}: {username}\n\n"
            f"{LANG_TEXT[lang]['total_work']}: {fmt_minutes(summary['total_work'])}\n"
            f"{LANG_TEXT[lang]['total_break']}: {fmt_minutes(summary['total_break'])}\n"
            f"{LANG_TEXT[lang]['leave_times']}: {summary['total_leave_times']}\n\n"
            f"{LANG_TEXT[lang]['meal']}: {summary['counts']['meal']} ({fmt_minutes(summary['durations']['meal'])})\n"
            f"{LANG_TEXT[lang]['toilet']}: {summary['counts']['toilet_small'] + summary['counts']['toilet_big']} ({fmt_minutes(summary['durations']['toilet_small'] + summary['durations']['toilet_big'])})\n"
            f"{LANG_TEXT[lang]['smoke']}: {summary['counts']['smoke']}\n"
        )

    await message.reply(text, parse_mode="HTML", reply_markup=get_menu(lang))

@dp.message(F.text.func(lambda s: text_in_keys(s, "leaderboard")))
async def cmd_leaderboard(message: types.Message):
    lang = detect_lang(message.from_user)
    chat_id = message.chat.id
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT DISTINCT user_id FROM work_sessions WHERE chat_id = ?", (chat_id,))
        rows = await cur.fetchall()
    users = [r[0] for r in rows]
    today = today_local_date()
    entries = []
    for uid in users:
        works, breaks = await get_day_intervals_for_user_in_chat(uid, chat_id, today)
        total_work = sum(minutes_between(s, e or now_utc()) for s, e in works if s)
        total_break = sum(minutes_between(s, e or now_utc()) for _, s, e in breaks if s)
        entries.append((uid, total_work - total_break, total_break))
    entries.sort(key=lambda x: x[1], reverse=True)
    lines = [f"{LANG_TEXT[lang]['leaderboard_title']}（{today.isoformat()}）"]
    if not entries:
        lines.append(LANG_TEXT[lang]["no_data"])
    else:
        pos = 1
        for uid, net_m, break_m in entries[:10]:
            try:
                member = await bot.get_chat_member(chat_id, uid)
                name = member.user.full_name or member.user.username or str(uid)
            except:
                name = str(uid)
            if lang == "zh":
                lines.append(f"{pos}. {name} — 工作 {fmt_minutes(net_m)}，休息 {fmt_minutes(break_m)}")
            elif lang == "en":
                lines.append(f"{pos}. {name} — Work {fmt_minutes(net_m)}, Break {fmt_minutes(break_m)}")
            else:
                lines.append(f"{pos}. {name} — Kerja {fmt_minutes(net_m)}, Istirahat {fmt_minutes(break_m)}")
            pos += 1
    await message.reply("\n".join(lines), reply_markup=get_menu(lang))

# ---------------------------
# 管理面板（多管理员） + 管理日志写入
# ---------------------------
def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

@dp.message(F.text.func(lambda s: text_in_keys(s, "settings")))
async def handler_settings(message: types.Message):
    lang = detect_lang(message.from_user)
    if not is_admin(message.from_user.id):
        await message.reply(LANG_TEXT[lang]["admin_only"])
        return
    await message.reply(LANG_TEXT[lang]["admin_menu_title"], reply_markup=get_admin_menu(lang))

@dp.callback_query(F.data == "admin:set_text")
async def admin_set_text(call: types.CallbackQuery):
    lang = detect_lang(call.from_user)
    if not is_admin(call.from_user.id):
        return await call.answer(LANG_TEXT[lang]["no_permission"], show_alert=True)
    await call.message.answer(LANG_TEXT[lang]["enter_new_text"])
    pending_media_for_chat[call.message.chat.id] = "awaiting_text"

@dp.message(F.text & (F.chat.id.in_(pending_media_for_chat.keys())))
async def handle_admin_input(message: types.Message):
    lang = detect_lang(message.from_user)
    chat_id = message.chat.id
    state = pending_media_for_chat.get(chat_id)
    if state == "awaiting_text":
        await set_chat_setting(chat_id, "reminder_text", message.text)
        await log_admin_action(chat_id, message.from_user.id, "set_reminder_text", message.text[:400])
        del pending_media_for_chat[chat_id]
        await message.reply(LANG_TEXT[lang]["text_updated"])
    elif state == "awaiting_media":
        await message.reply(LANG_TEXT[lang]["send_image"])

@dp.callback_query(F.data == "admin:set_media")
async def admin_set_media(call: types.CallbackQuery):
    lang = detect_lang(call.from_user)
    if not is_admin(call.from_user.id):
        return await call.answer(LANG_TEXT[lang]["no_permission"], show_alert=True)
    await call.message.answer(LANG_TEXT[lang]["send_image"])
    pending_media_for_chat[call.message.chat.id] = "awaiting_media"

@dp.message(F.photo)
async def handle_admin_photo(message: types.Message):
    lang = detect_lang(message.from_user)
    chat_id = message.chat.id
    if pending_media_for_chat.get(chat_id) == "awaiting_media":
        file_id = message.photo[-1].file_id
        await set_chat_setting(chat_id, "reminder_media_file_id", file_id)
        await log_admin_action(chat_id, message.from_user.id, "set_reminder_media", f"file_id:{file_id}")
        del pending_media_for_chat[chat_id]
        await message.reply(LANG_TEXT[lang]["image_updated"])

@dp.callback_query(F.data == "admin:toggle_weekly")
async def admin_toggle_weekly(call: types.CallbackQuery):
    lang = detect_lang(call.from_user)
    if not is_admin(call.from_user.id):
        return await call.answer(LANG_TEXT[lang]["no_permission"], show_alert=True)
    settings = await get_chat_settings(call.message.chat.id)
    new_value = 0 if settings["weekly_report_enabled"] else 1
    await set_chat_setting(call.message.chat.id, "weekly_report_enabled", new_value)
    await log_admin_action(call.message.chat.id, call.from_user.id, "toggle_weekly", f"set_to:{new_value}")
    status_text = LANG_TEXT[lang]["weekly_on"] if new_value else LANG_TEXT[lang]["weekly_off"]
    await call.message.edit_text(status_text, reply_markup=get_admin_menu(lang))

@dp.callback_query(F.data == "admin:toggle_monthly")
async def admin_toggle_monthly(call: types.CallbackQuery):
    lang = detect_lang(call.from_user)
    if not is_admin(call.from_user.id):
        return await call.answer(LANG_TEXT[lang]["no_permission"], show_alert=True)
    settings = await get_chat_settings(call.message.chat.id)
    new_value = 0 if settings["monthly_report_enabled"] else 1
    await set_chat_setting(call.message.chat.id, "monthly_report_enabled", new_value)
    await log_admin_action(call.message.chat.id, call.from_user.id, "toggle_monthly", f"set_to:{new_value}")
    status_text = LANG_TEXT[lang]["monthly_on"] if new_value else LANG_TEXT[lang]["monthly_off"]
    await call.message.edit_text(status_text, reply_markup=get_admin_menu(lang))

@dp.callback_query(F.data == "admin:reset_leaderboard")
async def admin_reset_leaderboard(call: types.CallbackQuery):
    lang = detect_lang(call.from_user)
    if not is_admin(call.from_user.id):
        return await call.answer(LANG_TEXT[lang]["no_permission"], show_alert=True)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM work_sessions WHERE chat_id = ?", (call.message.chat.id,))
        await db.execute("DELETE FROM break_sessions WHERE chat_id = ?", (call.message.chat.id,))
        await db.commit()
    await log_admin_action(call.message.chat.id, call.from_user.id, "reset_leaderboard", "cleared work_sessions and break_sessions")
    await call.message.answer(LANG_TEXT[lang]["reset_done"])
    await call.message.edit_text(LANG_TEXT[lang]["done"], reply_markup=get_admin_menu(lang))

@dp.callback_query(F.data == "admin:send_daily_report")
async def admin_send_daily_report(call: types.CallbackQuery):
    lang = detect_lang(call.from_user)
    if not is_admin(call.from_user.id):
        return await call.answer(LANG_TEXT[lang]["no_permission"], show_alert=True)
    today = today_local_date()
    chat_id = call.message.chat.id
    await send_report_for_chat(chat_id, "daily", today)
    await log_admin_action(chat_id, call.from_user.id, "manual_send_daily", f"sent daily for {today.isoformat()}")
    await call.message.answer(LANG_TEXT[lang]["daily_sent"])

# ---------------------------
# 超时提醒 watcher
# ---------------------------
async def break_overtime_watcher(user_id: int, chat_id: int, btype: str, start_dt_utc: datetime, lang_hint: str):
    limit_minutes = BREAK_LIMITS.get(btype, 5)
    limit_dt = start_dt_utc + timedelta(minutes=limit_minutes)
    while True:
        await asyncio.sleep(OVERTIME_REMINDER_INTERVAL * 60)
        now = now_utc()
        async with aiosqlite.connect(DB_PATH) as db:
            cur = await db.execute("SELECT id FROM break_sessions WHERE user_id=? AND chat_id=? AND end_time IS NULL", (user_id, chat_id))
            row = await cur.fetchone()
        if not row:
            break
        if now >= limit_dt:
            # 尝试获取用户语言
            try:
                member = await bot.get_chat_member(chat_id, user_id)
                lang = detect_lang(member.user)
            except:
                lang = lang_hint or "en"
            settings = await get_chat_settings(chat_id)
            default_text = LANG_TEXT[lang]["overtime_default"].format(uid=user_id)
            rtext = settings.get("reminder_text") or default_text
            try:
                media_file = settings.get("reminder_media_file_id")
                if media_file:
                    try:
                        await bot.send_photo(chat_id, media_file, caption=rtext, parse_mode="HTML")
                    except:
                        await bot.send_message(chat_id, rtext, parse_mode="HTML")
                else:
                    await bot.send_message(chat_id, rtext, parse_mode="HTML")
            except Exception as e:
                logger.exception(f"发送超时提醒失败: {e}")
            break  # 一次提醒后停止本次 watcher

# ---------------------------
# 报表：收集 / 生成 / 发送（Excel）
# ---------------------------
def safe_filename(s: str) -> str:
    # 移除文件名非法字符
    return re.sub(r'[\\/:"*?<>|]+', "_", s)

async def gather_users_in_chat(chat_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT DISTINCT user_id FROM work_sessions WHERE chat_id = ?", (chat_id,))
        rows = await cur.fetchall()
    return [r[0] for r in rows]

async def get_work_range_for_user(user_id: int, chat_id: int, start_utc: datetime, end_utc: datetime):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT start_time, end_time FROM work_sessions WHERE user_id=? AND chat_id=? AND start_time <= ? AND (end_time IS NULL OR end_time >= ?)",
            (user_id, chat_id, to_str(end_utc), to_str(start_utc))
        )
        rows = await cur.fetchall()
    starts = []
    ends = []
    total_work = 0
    for s, e in rows:
        ps = parse_str(s)
        pe = parse_str(e) if e else None
        if ps:
            starts.append(ps)
        if pe:
            ends.append(pe)
        total_work += minutes_between(ps, pe or end_utc)
    first_start = min(starts) if starts else None
    last_end = max(ends) if ends else None
    return first_start, last_end, total_work

async def get_break_summary_for_user(user_id: int, chat_id: int, start_utc: datetime, end_utc: datetime):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT type, start_time, end_time FROM break_sessions WHERE user_id=? AND chat_id=? AND start_time <= ? AND (end_time IS NULL OR end_time >= ?)",
            (user_id, chat_id, to_str(end_utc), to_str(start_utc))
        )
        rows = await cur.fetchall()
    total_break = 0
    leave_count = 0
    for t, s, e in rows:
        ps = parse_str(s)
        pe = parse_str(e) if e else end_utc
        if ps:
            total_break += minutes_between(ps, pe)
            leave_count += 1
    return total_break, leave_count

async def send_report_for_chat(chat_id: int, period: str, base_date: date):
    # 计算 local_start / local_end
    if period == "daily":
        local_start = datetime.combine(base_date, time.min)
        local_end = datetime.combine(base_date, time.max)
        prefix = "日报"
    elif period == "weekly":
        start_local = base_date - timedelta(days=base_date.weekday())
        local_start = datetime.combine(start_local, time.min)
        local_end = local_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
        prefix = "周报"
    elif period == "monthly":
        start_local = base_date.replace(day=1)
        if start_local.month == 12:
            next_month = start_local.replace(year=start_local.year + 1, month=1, day=1)
        else:
            next_month = start_local.replace(month=start_local.month + 1, day=1)
        local_start = datetime.combine(start_local, time.min)
        local_end = datetime.combine(next_month - timedelta(seconds=1), time.max)
        prefix = "月报"
    else:
        return

    utc_start = local_start - LOCAL_OFFSET
    utc_end = local_end - LOCAL_OFFSET

    users = await gather_users_in_chat(chat_id)
    if not users:
        logger.info(f"chat {chat_id} 没有用户数据，跳过 {period} 报表。")
        return

    rows = []
    for uid in users:
        try:
            member = await bot.get_chat_member(chat_id, uid)
            name = member.user.full_name or member.user.username or str(uid)
        except:
            name = str(uid)
        first_start, last_end, total_work = await get_work_range_for_user(uid, chat_id, utc_start, utc_end)
        total_break, leave_count = await get_break_summary_for_user(uid, chat_id, utc_start, utc_end)
        first_start_s = fmt_hm_local(first_start) if first_start else "-"
        last_end_s = fmt_hm_local(last_end) if last_end else "-"
        rows.append((name, first_start_s, last_end_s, total_work, total_break, leave_count))

    rows.sort(key=lambda x: x[3], reverse=True)

    # 生成 Excel 报表
    wb = Workbook()
    ws = wb.active
    ws.title = f"{prefix}"

    headers = ["姓名", "上班时间", "下班时间", "工作时间(文本)", "休息时间(文本)", "离开次数", "工作时间(分钟)", "休息时间(分钟)"]
    ws.append(headers)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for name, start_s, end_s, work_m, break_m, leave_cnt in rows:
        ws.append([
            name,
            start_s,
            end_s,
            fmt_minutes(work_m),
            fmt_minutes(break_m),
            leave_cnt,
            work_m,
            break_m
        ])

    # 自动列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    # 保存到内存
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    bytes_data = file_bytes.getvalue()

    # 获取群名
    try:
        chat = await bot.get_chat(chat_id)
        chat_title = chat.title or "群名未知"
    except Exception:
        chat_title = "群名未知"

    fname_safe = safe_filename(f"{prefix}_{chat_title}_{base_date.isoformat()}.xlsx")
    tz_hour = int(LOCAL_OFFSET.total_seconds() // 3600)
    caption = f"📤 [{chat_title}] (ID: {chat_id}) 的 {prefix}\n{LANG_TEXT['zh']['tz_label']}：UTC{tz_hour:+d}"

    # 发送给所有管理员
    for admin in ADMIN_IDS:
        try:
            buffered = BufferedInputFile(bytes_data, filename=fname_safe)
            await bot.send_document(admin, document=buffered, caption=caption)
            logger.info(f"✅ 已发送 {prefix} 给管理员 {admin}")
        except Exception as e:
            logger.warning(f"发送报表给管理员 {admin} 失败: {e}")

# ---------------------------
# 定时任务（apscheduler）
# ---------------------------
scheduler = AsyncIOScheduler(timezone="Asia/Jakarta")

@scheduler.scheduled_job(CronTrigger(hour=DAILY_REPORT_HOUR, minute=0))
async def scheduled_daily_report():
    today = today_local_date()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT DISTINCT chat_id FROM work_sessions")
        rows = await cur.fetchall()
    for (chat_id,) in rows:
        await send_report_for_chat(chat_id, "daily", today)

@scheduler.scheduled_job(CronTrigger(day_of_week="mon", hour=WEEKLY_REPORT_HOUR, minute=0))
async def scheduled_weekly_report():
    today = today_local_date()
    chats = await get_chats_with_setting_enabled("weekly_report_enabled")
    for cid in chats:
        await send_report_for_chat(cid, "weekly", today)

@scheduler.scheduled_job(CronTrigger(day=MONTHLY_REPORT_DAY, hour=MONTHLY_REPORT_HOUR, minute=0))
async def scheduled_monthly_report():
    today = today_local_date()
    chats = await get_chats_with_setting_enabled("monthly_report_enabled")
    for cid in chats:
        await send_report_for_chat(cid, "monthly", today)

# 手动触发日报命令（管理员）—— 同步三语反馈
@dp.message(F.text.func(lambda s: ("手动发送日报" in s) or ("Send Daily Report" in s) or ("Kirim Laporan Harian" in s)))
async def manual_daily_report(message: types.Message):
    lang = detect_lang(message.from_user)
    if message.from_user.id not in ADMIN_IDS:
        await message.reply(LANG_TEXT[lang]["not_admin"])
        return
    today = today_local_date()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT DISTINCT chat_id FROM work_sessions")
        rows = await cur.fetchall()
    for (chat_id,) in rows:
        await send_report_for_chat(chat_id, "daily", today)
    await message.reply(LANG_TEXT[lang]["manual_daily_done"])

# ---------------------------
# 启动
# ---------------------------
async def main():
    await init_db()
    scheduler.start()
    logger.info("调度器已启动（日报/周报/月报）。")
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("已停止。")
